from __future__ import annotations

# from os.path import join
# import datetime
from openpyxl import Workbook
import pandas as pd
import geopandas as gpd

from abc import ABC, abstractmethod
# from os import walk
from typing import Any, Dict, Iterator, List, Optional

from flask import current_app

from openpyxl.styles import (
    NamedStyle,
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    Color
)

# from qhawariy.utilities import builtins
from qhawariy.utilities.helpers import a_dict

# Estilo para las hojas de excel
RED = Color(rgb='00EB0046')
BLUE = Color(rgb='000039A6')
GREEN = Color(rgb='0005BE50')
GRAY = Color(rgb='00878C8F')
GRAY_2 = Color(rgb='00E9E9E9')
GRAY_3 = Color(rgb='004D4D4D')
BLACK = Color(rgb='000F191E')

FONT = Font(
    name='Calibri',
    size=12,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='000000'
)
FONT_BOLD = Font(
    name='Calibri',
    size=12,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=GRAY_3
)
FONT_GRAY = Font(
    name='Calibri',
    size=12,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=GRAY_3
)
FONT_TITLE = Font(
    name='Calibri',
    size=20,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=BLACK
)
FONT_SMALL = Font(
    name='Calibri',
    size=9,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=BLACK
)
FONT_WHITE = Font(
    name='Calibri',
    size=12,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FFFFFF'
)
FONT_BLACK = Font(
    name='Calibri',
    size=12,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color=BLACK
)

FILL_RED = PatternFill("solid", fgColor=RED)
FILL_BLUE = PatternFill("solid", fgColor=BLUE)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_GRAY_2 = PatternFill("solid", fgColor=GRAY_2)

BORDER = Border(
    left=Side(style='thin', color=GRAY),
    right=Side(style='thin', color=GRAY),
    top=Side(style='thin', color=GRAY),
    bottom=Side(style='thin', color=GRAY)
)

ALIGNMENT = Alignment(
    horizontal='center',
    vertical='center',
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0
)
ALIGNMENT_JUSTIFY = Alignment(
    horizontal='distributed',
    vertical='center',
    text_rotation=0,
    wrap_text=False,
    justifyLastLine=True,
    shrink_to_fit=False,
    indent=0
)
ALIGNMENT_JUSTIFY_TEXT = Alignment(
    horizontal='distributed',
    vertical='distributed',
    text_rotation=0,
    justifyLastLine=True,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0
)

NUMBER_FORMAT = 'general'
NUMBER = '0'
TIME = 'hh:mm:ss AM/PM'

PROTECTION = Protection(locked=True, hidden=False)


# Clases para crear diferentes formas y dar formatos a archivos
class Contexto():
    """
    Class Contexto: Define la interface de interes del cliente.
    """
    def __init__(self, estrategia: Estrategia) -> None:
        self._estrategia = estrategia

    @property
    def estrategia(self) -> Estrategia:
        return self._estrategia

    @estrategia.setter
    def estrategia(self, estrategia: Estrategia) -> None:
        self._estrategia = estrategia

    def realizar(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        routes: Optional[List[str]],
        sheetnames: Optional[List[str]],
        date: Optional[str]
    ) -> None:
        # Construir archivo usando la estrategia
        self._estrategia.construirArchivo(data, filename, routes, sheetnames, date)


class Estrategia(ABC):
    """
    Clase Estrategia: Declara operaciones comunes para todo las versiones soportadas
    de algun algoritmo.

    El Contexto usa esta interace para llamar al algoritmo definido por un conreto
    estrategia.
    """

    @abstractmethod
    def construirArchivo(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        routes: Optional[List[str]],
        sheetnames: Optional[List[str]],
        date: Optional[str]
    ):
        pass


class EstrategiaExcelLista(Estrategia):
    def construirArchivo(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        routes: Optional[List[str]],
        sheetnames: Optional[List[str]],
        date: Optional[str]
    ):
        converted: Iterator[Optional[Dict[str, Any]]]
        converted = (a_dict(item) for item in data)
        data_clean: List[Dict[str, Any]] = [
            d for d in converted if d is not None  # type: ignore
        ]
        df: pd.DataFrame = pd.DataFrame(data_clean)

        if not sheetnames:
            sheetnames = ["Sheet1"]

        writer: pd.ExcelWriter = pd.ExcelWriter(filename, engine='openpyxl')
        try:
            for name in sheetnames:
                df.to_excel(  # type: ignore
                    writer,
                    startrow=0,
                    merge_cells=False,
                    sheet_name=name
                )
        finally:
            writer.close()


class EstrategiaExcelResumen(Estrategia):
    # Inicializar estilos
    estilo = NamedStyle(name="estilo")
    estilo.font = FONT_BLACK
    estilo.border = BORDER
    estilo.alignment = ALIGNMENT

    head = NamedStyle(name="head")
    head.font = FONT_WHITE
    head.fill = FILL_GREEN
    head.border = BORDER
    head.alignment = ALIGNMENT_JUSTIFY_TEXT
    head.number_format = NUMBER_FORMAT

    titulo = NamedStyle(name="titulo")
    titulo.font = FONT_TITLE
    titulo.alignment = ALIGNMENT

    aviso = NamedStyle(name='aviso')
    aviso.font = FONT_SMALL
    aviso.number_format = NUMBER_FORMAT
    aviso.alignment = ALIGNMENT_JUSTIFY

    # def construirArchivo(
    # self, data: List, filename: str, sheetnames: List | None, date: str | None
    # ):
    def construirArchivo(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        routes: Optional[List[str]],
        sheetnames: Optional[List[str]],
        date: Optional[str]
    ):
        if sheetnames is None:
            # Crear nombres por defecto
            sheetnames = [f"Sheet{i+1}" for i in range(len(data))]
        else:
            # Eliminar caracteres no permitidos
            sheetnames = [str(s).replace("/", "") for s in sheetnames]
            if len(sheetnames) < len(data):
                # Rellenar los nombres por defecto
                extra = [f"Sheet{i+1}" for i in range(len(sheetnames), len(data))]
                sheetnames.extend(extra)
        if routes is None:
            routes = [""] * len(data)
        else:
            # asegurar que routes tenga al menos la longitud necesaria
            if len(routes) < len(data):
                routes = list(routes) + [""] * (len(data) - len(routes))

        date_text = date or ""

        inicia_fila = 4
        # Escribir los df en diferentes hojas de excel usando context manager
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for idx, df in enumerate(data):
                sheet_name = str(sheetnames[idx])
                df.to_excel(  # type: ignore
                    writer,
                    startrow=inicia_fila,
                    merge_cells=False,
                    sheet_name=sheet_name
                )
            # Obtener workbook y recorrer hojas con enumerate para índice seguro
            workbooks = writer.book  # type: ignore
            for count, worksheet in enumerate(workbooks.worksheets):  # type: ignore
                # Formato y contenido de cabecera
                worksheet.merge_cells("A1:H1")
                worksheet.merge_cells("B2:F2")
                worksheet.merge_cells("A3:H3")

                worksheet.sheet_format.baseColWidth = 9

                worksheet["A1"] = "Nombre de la empresa"
                worksheet["A2"] = "Recorrido:"
                worksheet["B2"] = routes[count]
                worksheet["G2"] = "Versión"
                worksheet["H2"] = "v.1.0"
                worksheet["A3"] = "Resumen semanal " + date_text

                worksheet.cell(1, 1).style = self.titulo

                for row in worksheet["A2:H3"]:
                    for cell in row:
                        cell.font = FONT_GRAY
                worksheet.cell(2, 1).font = FONT_BOLD
                worksheet.cell(2, 7).font = FONT_BOLD

                # Formatos de tabla
                for row in worksheet[f"A{inicia_fila+1}:H{inicia_fila+1}"]:
                    for cell in row:
                        cell.style = self.head

                celdas_A_H = f"A{inicia_fila+2}:H{worksheet.max_row}"
                for row in worksheet[celdas_A_H]:
                    for cell in row:
                        cell.style = self.estilo

                celdas_A_A = f"A{inicia_fila+2}:A{worksheet.max_row}"
                for row in worksheet[celdas_A_A]:
                    for cell in row:
                        cell.number_format = TIME
                        cell.fill = FILL_GRAY_2

                celdas_B_H = f"B{inicia_fila+2}:H{worksheet.max_row}"
                for row in worksheet[celdas_B_H]:
                    for cell in row:
                        cell.number_format = NUMBER

                # Footer / aviso de copyright
                ultimo_cell = worksheet.max_row + 2
                footer_aviso = worksheet.cell(ultimo_cell, 1)
                worksheet.merge_cells(f"A{ultimo_cell}:H{ultimo_cell+1}")
                footer_aviso.value = (
                    "* Este modelo de formato de lista de programacion,"
                    " fue elaborado por SEQhawariy.\n Queda prohibido su uso,"
                    " comercialización y/o difusión, sin previa autorización o permiso"
                    " de SEQhawariy"
                )  # type: ignore
                footer_aviso.style = self.aviso

        # # Para los nombres
        # for i in range(len(sheetnames)):
        #     # eliminar los caracteres no permitidos por titulo de hojas de excel
        #     sheetnames[i] = str(sheetnames[i]).replace("/", "")

        # # Numero de columna en la que se escribira los datos
        # inicia_fila = 4

        # # Escribir los df en diferente hojas de excel
        # writer = pd.ExcelWriter(filename, engine='openpyxl')
        # for df in range(len(data)):
        #     sheet_name = str(sheetnames[df])
        #     data[df].to_excel(  # type: ignore
        #         writer,
        #         startrow=inicia_fila,
        #         merge_cells=False,
        #         sheet_name=sheet_name
        #     )

        # # Obtener objetos del libro y hoja de trabajo
        # workbooks = writer.book  # type: ignore
        # count = 0
        # for worksheet in workbooks:  # type: ignore
        #     # Combinar celdas para formato de titulo
        #     worksheet.merge_cells('A1:H1')  # type: ignore
        #     worksheet.merge_cells('B2:F2')  # type: ignore
        #     worksheet.merge_cells('A3:H3')  # type: ignore

        #     worksheet.sheet_format.baseColWidth = 9  # type: ignore

        #     # Agregando contenido a la cabecera
        #     worksheet['A1'] = "Nombre de la empresa"
        #     worksheet['A2'] = "Recorrido:"
        #     worksheet['B2'] = routes[count]
        #     # worksheet['B2']=worksheet.title.title()
        #     worksheet['G2'] = "Versión"
        #     worksheet['H2'] = "v.1.0"
        #     worksheet['A3'] = "Resumen semanal " + date_text

        #     worksheet.cell(1, 1).style = self.titulo  # type: ignore
        #     for row in worksheet['A{i}:H{f}'.format(i=2, f=3)]:  # type: ignore
        #         for cell in row:  # type: ignore
        #             cell.font = FONT_GRAY
        #     worksheet.cell(2, 1).font = FONT_BOLD  # type: ignore
        #     worksheet.cell(2, 7).font = FONT_BOLD  # type: ignore

        #     # Agregando un formato para la cabecera cuadro
        #     for row in worksheet['A{i}:H{i}'.format(i=inicia_fila+1)]:  # type: ignore
        #         for cell in row:  # type: ignore
        #             cell.style = self.head
        #     # Agregando un formato para la  cuerpo del cuadro
        #     celdas_A_H = 'A{co}:H{c}'.format(
        #         co=inicia_fila+2,
        #         c=worksheet.max_row  # type: ignore
        #     )
        #     for row in worksheet[celdas_A_H]:  # type: ignore
        #         for cell in row:  # type: ignore
        #             cell.style = self.estilo

        #     celdas_A_A = 'A{co}:A{c}'.format(
        #         co=inicia_fila+2,
        #         c=worksheet.max_row  # type: ignore
        #     )
        #     for row in worksheet[celdas_A_A]:  # type: ignore
        #         for cell in row:  # type: ignore
        #             cell.number_format = TIME
        #             cell.fill = FILL_GRAY_2

        #     celdas_B_H = 'B{co}:H{c}'.format(co=inicia_fila+2, c=worksheet.max_row)
        #     for row in worksheet[celdas_B_H]:
        #         for cell in row:
        #             cell.number_format = NUMBER

        #     # Para dar formato al aviso de derecho de autor
        #     ultimo_cell = worksheet.max_row+2
        #     worksheet.merge_cells("A{u}:H{uf}".format(
        #         u=ultimo_cell,
        #         uf=ultimo_cell+1
        #     ))
        #     footer_aviso = worksheet.cell(ultimo_cell, 1)
        #     footer_aviso.value = '* Este modelo de formato de lista de programacion,'\
        #         'fue elaborado por SEQhawariy.\n Queda prohibido su uso,'\
        #         'comercialización y/o difusión, sin previa autorización o permiso de'\
        #         'SEQhawariy'
        #     footer_aviso.style = self.aviso

            # count = count+1
                count = count+1

        # Cerrar el escritor Excel de pandas
        writer.close()


class Archivo(ABC):
    """
    Clase Archivo: que implementa acciones para
    administrar un archivo
    """

    @abstractmethod
    def guardar(self):
        pass


class ArchivoExcel(Archivo):
    """
    Clase ArchivoExcel: Permite manejar un archivo excel
    """
    def __init__(
        self,
        nombre_archivo: str,
        dataframe: list,
        tipo: str,
        nombres_hojas: list | None,
        recorridos: list | None,
        fecha: str | None
    ):
        self.nombre = nombre_archivo
        self.dataframe = dataframe
        self.fecha = fecha
        self.recorridos = recorridos
        self.nombres_hojas = nombres_hojas
        self.tipo = tipo
        fname = current_app.config['DOWNLOAD_FOLDER']+"\\"+nombre_archivo+".xlsx"
        self.ruta_archivo = fname

    def guardar(self):
        contexto = None
        if self.tipo == 'lista':
            contexto = Contexto(EstrategiaExcelLista())
        elif self.tipo == 'resumen':
            contexto = Contexto(EstrategiaExcelResumen())

        contexto.realizar(data=self.dataframe,
                          filename=self.ruta_archivo,
                          routes=self.recorridos,
                          sheetnames=self.nombres_hojas,
                          date=self.fecha
                          )


class ArchivoCSV(Archivo):
    def __init__(self):
        self.nombre = 'csv'

    def guardar(self):
        print(f'guardado {self.nombre}')


class ArchivoShapefile(Archivo):
    """
    Clase ArchivoShapefile: Permite administrar el archivo shapefile
    parameter: nombre_archivo: Debe contener la extesion de archivo ".shp"
    """

    def __init__(self, filename):
        self.nombre_archivo = filename
        ruta_archivo = current_app.config['SHAPEFILE_FOLDER']+"\\"+self.nombre_archivo
        self.geo_data_frame = gpd.read_file(ruta_archivo)

    def guardar(self):
        pass

    @staticmethod
    def obtenerLongitudLatitud(frame):
        """
        funcion auxiliar que retorna longitud y latitud de una DataFrame de Geopandas
        """
        xy = frame.geometry.xy
        longs = xy[0].tolist()
        lats = xy[1].tolist()
        return [list(z) for z in zip(lats, longs)]


class Factory(ABC):
    def __init__(self):
        pass

    def crearArchivo(self) -> Archivo:
        pass


class FactoryExcel(Factory):
    def crearArchivo(
        self,
        filename: str,
        dataframe: pd.DataFrame,
        tipo: str,
        sheetnames: List,
        recorrido: List,
        date: str
    ):
        archivo_excel = ArchivoExcel(
            nombre_archivo=filename,
            tipo=tipo,
            dataframe=dataframe,
            nombres_hojas=sheetnames,
            recorridos=recorrido,
            fecha=date
        )
        return archivo_excel


class FactoryCSV(Factory):
    def crearArchivo(self):
        archivo_csv = ArchivoCSV()
        return archivo_csv


class FactoryShapefile(Factory):
    def crearArchivo(self, filename: str):
        archivo_shapefile = ArchivoShapefile(filename=filename)
        return archivo_shapefile
